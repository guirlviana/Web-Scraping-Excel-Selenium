from selenium import webdriver
import os
import openpyxl
import time

class BuscarProduto:
    def __init__(self):
        self.driver = webdriver.Chrome(executable_path= os.getcwd() + os.sep + 'chromedriver.exe')

    def Iniciar(self):
        self.produto = input('Qual produto buscar?')
        self.FazerPlanilha()
        self.EncontrarElementos()
    
    def FazerPlanilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Valores')
        self.planilha_valores = self.planilha['Valores']
        self.planilha_valores.cell(row= 1, column=1, value='Titulo')
        self.planilha_valores.cell(row= 1, column=2, value='Localizacao')
        self.planilha_valores.cell(row= 1, column=3, value='Precos')
        
    def EncontrarElementos(self):
        try:
            for self.numPags in range(1,3):
                self.IrproxPag()
                self.titulo = self.driver.find_elements_by_xpath('//h2[@class="fnmrjs-10 deEIZJ"]')
                self.localizacao = self.driver.find_elements_by_xpath('//p[@class="fnmrjs-13 hdwqVC"]')
                self.precos = self.driver.find_elements_by_xpath('//p[@class="fnmrjs-16 jqSHIm"]')
                self.ArmazenarValores()        
        except Exception as erro:
            print('Fim')
    
    def ArmazenarValores(self):
        for indice in range(0, len(self.titulo)):
            nova_linha = [self.titulo[indice].text, 
            self.localizacao[indice].text, self.precos[indice].text]
            self.planilha_valores.append(nova_linha)
        self.planilha.save(f'Preços {self.produto}.xlsx')
    
    def IrproxPag(self):
        time.sleep(2) 
        self.driver.get(f'https://sp.olx.com.br/sao-paulo-e-regiao/zona-leste?o={self.numPags}&q={self.produto}')
        

root = BuscarProduto()
root.Iniciar()